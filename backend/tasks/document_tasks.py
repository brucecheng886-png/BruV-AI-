"""
文件攝取 Celery 任務

Pipeline：
  MinIO 下載 → 解析 → SentenceWindow 分塊 (Excel 除外)
  → bge-m3 嵌入 → Qdrant 寫入 → LLM 實體分析 → Neo4j 寫入
  → PG chunks 寫入 → 更新文件狀態

所有跨庫操作由 SagaLog 保護（必要元件）。
"""
import io
import json
import logging
import re
import uuid
from typing import List, Tuple

import httpx
import psycopg2
from celery import shared_task
from celery.utils.log import get_task_logger
from minio import Minio
from neo4j import GraphDatabase
from qdrant_client import QdrantClient
from qdrant_client.models import PointStruct

logger = get_task_logger(__name__)


# ── 連線工廠（Task 內建立，不共用）────────────────────────────

def _settings():
    from config import settings
    return settings


def _pg_conn():
    s = _settings()
    return psycopg2.connect(
        host=s.POSTGRES_HOST, port=s.POSTGRES_PORT,
        database=s.POSTGRES_DB, user=s.POSTGRES_USER,
        password=s.POSTGRES_PASSWORD,
    )


def _minio():
    s = _settings()
    return Minio(s.MINIO_ENDPOINT,
                 access_key=s.MINIO_ACCESS_KEY,
                 secret_key=s.MINIO_SECRET_KEY,
                 secure=False)


def _qdrant():
    s = _settings()
    return QdrantClient(host=s.QDRANT_HOST, port=s.QDRANT_PORT, api_key=s.QDRANT_API_KEY or None, https=False)


def _neo4j():
    s = _settings()
    return GraphDatabase.driver(s.NEO4J_URI, auth=(s.NEO4J_USER, s.NEO4J_PASSWORD))


# ── 文件解析 ───────────────────────────────────────────────────

def _parse_pdf(data: bytes) -> List[Tuple[str, int]]:
    from pypdf import PdfReader
    reader = PdfReader(io.BytesIO(data))
    return [(p.extract_text() or "", i + 1)
            for i, p in enumerate(reader.pages)
            if (p.extract_text() or "").strip()]


def _parse_docx(data: bytes) -> List[Tuple[str, int]]:
    from docx import Document as DocxDoc
    doc = DocxDoc(io.BytesIO(data))
    text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    return [(text, 1)]


def _parse_xlsx(data: bytes) -> List[Tuple[str, int]]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    results = []
    for sheet_idx, ws in enumerate(wb.worksheets, 1):
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h) if h is not None else f"col{i}"
                   for i, h in enumerate(rows[0])]
        for row in rows[1:]:
            if all(v is None for v in row):
                continue
            parts = [f"{h}: {v}" for h, v in zip(headers, row) if v is not None]
            text = " | ".join(parts)
            if text.strip():
                results.append((text, sheet_idx))
    return results


def _parse_html(data: bytes) -> List[Tuple[str, int]]:
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(data, "html.parser")
    for tag in soup(["script", "style"]):
        tag.decompose()
    return [(soup.get_text(separator="\n"), 1)]


def _parse_text(data: bytes) -> List[Tuple[str, int]]:
    # 嘗試 UTF-8（含 BOM）→ fallback CP950（繁中 Windows）
    for enc in ("utf-8-sig", "utf-8", "cp950", "latin-1"):
        try:
            text = data.decode(enc)
            return [(text, 1)]
        except (UnicodeDecodeError, LookupError):
            continue
    return [(data.decode("utf-8", errors="replace"), 1)]


_PARSERS = {
    "pdf": _parse_pdf,
    "docx": _parse_docx,
    "xlsx": _parse_xlsx,
    "html": _parse_html,
    "txt":  _parse_text,
    "md":   _parse_text,
    "csv":  _parse_text,
}


# ── 分塊 ──────────────────────────────────────────────────────

CHUNK_SIZE = 400     # 字元（軟上限，可被 DB 設定覆蓋）
WINDOW     = 3       # 句子視窗


def _get_chunk_size(pg) -> int:
    """從 system_settings 讀取 doc_chunk_size，失敗時回傳預設值"""
    try:
        cur = pg.cursor()
        cur.execute("SELECT value FROM system_settings WHERE key = 'doc_chunk_size'")
        row = cur.fetchone()
        cur.close()
        return int(row[0]) if row else CHUNK_SIZE
    except Exception:
        return CHUNK_SIZE


def _get_kb_config(pg, kb_id: str | None) -> dict:
    """讀取 KB 級設定（embedding_model / embedding_provider / chunk_size 等）

    若 KB 沒設或欄位為 NULL，對應值回傳 None，由呼叫端 fallback 至全域。
    雲端 provider 會額外從 llm_models 取出 base_url / 解密後的 api_key。
    """
    cfg: dict = {
        "embedding_model": None,
        "embedding_provider": None,
        "embedding_base_url": None,
        "embedding_api_key": None,
        "chunk_size": None,
        "chunk_overlap": None,
    }
    if not kb_id:
        return cfg
    try:
        cur = pg.cursor()
        cur.execute(
            "SELECT embedding_model, embedding_provider, chunk_size, chunk_overlap "
            "FROM knowledge_bases WHERE id=%s",
            (kb_id,),
        )
        row = cur.fetchone()
        cur.close()
        if not row:
            return cfg
        cfg["embedding_model"]    = row[0]
        cfg["embedding_provider"] = row[1]
        cfg["chunk_size"]         = row[2]
        cfg["chunk_overlap"]      = row[3]

        if cfg["embedding_model"] and cfg["embedding_provider"] and cfg["embedding_provider"] != "ollama":
            cur2 = pg.cursor()
            cur2.execute(
                "SELECT base_url, api_key FROM llm_models "
                "WHERE name=%s AND provider=%s LIMIT 1",
                (cfg["embedding_model"], cfg["embedding_provider"]),
            )
            r2 = cur2.fetchone()
            cur2.close()
            if r2:
                cfg["embedding_base_url"] = r2[0]
                if r2[1]:
                    try:
                        from utils.crypto import decrypt_secret
                        cfg["embedding_api_key"] = decrypt_secret(r2[1])
                    except Exception:
                        cfg["embedding_api_key"] = None
    except Exception as exc:
        logger.warning("Failed to load KB config for %s: %s", kb_id, exc)
    return cfg


def _sentence_window_chunks(text: str, page: int, chunk_size: int = CHUNK_SIZE) -> List[dict]:
    """SentenceWindow 分塊：3 句滑動，附前後文 window_context"""
    sentences = [s.strip() for s in
                 re.split(r"(?<=[\u3002\uff01\uff1f.!?\n])\s*", text) if s.strip()]
    if not sentences:
        return []
    chunks = []
    i = 0
    while i < len(sentences):
        group = sentences[i: i + WINDOW]
        content = " ".join(group)
        if len(content) > chunk_size * 2:
            content = sentences[i]
        prev = sentences[i - 1] if i > 0 else ""
        nxt  = sentences[i + len(group)] if i + len(group) < len(sentences) else ""
        window_ctx = " ".join(filter(None, [prev, content, nxt]))
        chunks.append({
            "content": content,
            "window_context": window_ctx,
            "page_number": page,
        })
        i += max(1, WINDOW - 1)
    return chunks


def _xlsx_chunks(pages: List[Tuple[str, int]]) -> List[dict]:
    """Excel：每行直接成為一個 chunk（不做 sentence window）"""
    return [{"content": t, "window_context": t, "page_number": pg}
            for t, pg in pages if t.strip()]


# ── Ollama 工具 ───────────────────────────────────────────────

def _embed_texts(texts: List[str], model: str, base_url: str) -> List[List[float]]:
    """批次嵌入，每次最多 8 筆；若批次失敗則逐一降級；NaN 以 0.0 取代"""
    import math

    # 預先偵測向量維度（用一個短文字探測）
    _DIM_CACHE: List[int] = []
    def _get_dim() -> int:
        if _DIM_CACHE:
            return _DIM_CACHE[0]
        try:
            r = httpx.post(f"{base_url}/api/embed", json={"model": model, "input": ["dim"]}, timeout=30)
            if r.status_code == 200:
                vecs = r.json().get("embeddings", [[]])
                dim = len(vecs[0]) if vecs else 1024
            else:
                dim = 1024
        except Exception:
            dim = 1024
        _DIM_CACHE.append(dim)
        return dim

    def _clean_vec(vec: List[float]) -> List[float]:
        return [0.0 if (v is None or (isinstance(v, float) and math.isnan(v))) else v for v in vec]

    def _embed_one(text: str) -> List[float]:
        try:
            with httpx.Client(timeout=120) as client:
                r = client.post(f"{base_url}/api/embed", json={"model": model, "input": [text]})
                r.raise_for_status()
                vecs = r.json().get("embeddings", [[]])
                return _clean_vec(vecs[0]) if vecs else [0.0] * _get_dim()
        except Exception:
            logger.warning("Failed to embed text (fallback to zero): %s", text[:60])
            return [0.0] * _get_dim()

    all_vecs: List[List[float]] = []
    batch = 8
    with httpx.Client(timeout=120) as client:
        for i in range(0, len(texts), batch):
            batch_texts = texts[i: i + batch]
            try:
                resp = client.post(
                    f"{base_url}/api/embed",
                    json={"model": model, "input": batch_texts},
                )
                resp.raise_for_status()
                vecs = resp.json().get("embeddings", [])
                if len(vecs) == len(batch_texts):
                    all_vecs.extend([_clean_vec(v) for v in vecs])
                    continue
            except Exception:
                pass
            # 批次失敗 → 逐一嵌入
            logger.warning("Batch embed failed (batch %d-%d), falling back to individual", i, i + len(batch_texts))
            for t in batch_texts:
                all_vecs.append(_embed_one(t))
    return all_vecs


def _embed_texts_openai(texts: List[str], model: str, base_url: str, api_key: str) -> List[List[float]]:
    """OpenAI 相容 embedding API（POST {base}/v1/embeddings, Bearer auth）"""
    import math
    url = base_url.rstrip("/")
    if not url.endswith("/v1"):
        url = f"{url}/v1"
    url = f"{url}/embeddings"
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}

    def _clean(vec):
        return [0.0 if (v is None or (isinstance(v, float) and math.isnan(v))) else v for v in vec]

    out: List[List[float]] = []
    batch = 16
    with httpx.Client(timeout=120) as client:
        for i in range(0, len(texts), batch):
            chunk = texts[i: i + batch]
            try:
                r = client.post(url, headers=headers, json={"model": model, "input": chunk})
                r.raise_for_status()
                data = r.json().get("data", [])
                if len(data) == len(chunk):
                    out.extend([_clean(d.get("embedding", [])) for d in data])
                    continue
            except Exception as exc:
                logger.warning("OpenAI embed batch failed: %s", exc)
            # fallback: 逐一
            for t in chunk:
                try:
                    r = client.post(url, headers=headers, json={"model": model, "input": [t]})
                    r.raise_for_status()
                    d = r.json().get("data", [{}])[0]
                    out.append(_clean(d.get("embedding", [])))
                except Exception:
                    out.append([])
    return out


def _embed_dispatch(texts: List[str], kb_cfg: dict, doc_id: str | None = None) -> List[List[float]]:
    """依 KB 設定路由到 ollama 或雲端 embedding；C4：雲端失敗 → fallback ollama bge-m3 並寫入 documents.ingestion_warnings"""
    s = _settings()
    provider = (kb_cfg.get("embedding_provider") or "").lower()
    model    = kb_cfg.get("embedding_model")

    def _record_warning(reason: str):
        if not doc_id:
            return
        try:
            from prompts import EMBEDDING_FALLBACK_NOTICE
            with _pg_conn() as pg, pg.cursor() as cur:
                cur.execute(
                    "UPDATE documents SET ingestion_warnings = COALESCE(ingestion_warnings,'[]'::jsonb) || %s::jsonb "
                    "WHERE id=%s",
                    (json.dumps([{
                        "type": "embedding_fallback",
                        "reason": reason,
                        "notice": EMBEDDING_FALLBACK_NOTICE,
                    }], ensure_ascii=False), doc_id),
                )
                pg.commit()
        except Exception as _we:
            logger.warning("Failed to record ingestion_warning: %s", _we)

    if model and provider and provider != "ollama":
        base_url = kb_cfg.get("embedding_base_url") or "https://api.openai.com"
        api_key  = kb_cfg.get("embedding_api_key") or ""
        if not api_key:
            logger.warning("KB embedding api_key missing for provider=%s, fallback to ollama", provider)
            _record_warning(f"雲端 provider={provider} 未設定 api_key")
        else:
            try:
                vecs = _embed_texts_openai(texts, model, base_url, api_key)
                # 檢查是否全部為空向量（表示全部失敗）
                if vecs and any(v for v in vecs):
                    return vecs
                logger.warning("Cloud embedding returned all-empty vectors, falling back to ollama")
                _record_warning(f"雲端 provider={provider} model={model} 回傳空向量")
            except Exception as _ee:
                logger.error("Cloud embedding failed (%s), falling back to ollama bge-m3: %s", provider, _ee)
                _record_warning(f"雲端 provider={provider} model={model} 失敗：{_ee}")

    # ollama 分支：使用 KB 指定 model 或全域 fallback
    use_model = model if (model and (not provider or provider == "ollama")) else s.OLLAMA_EMBED_MODEL
    return _embed_texts(texts, use_model, s.OLLAMA_BASE_URL)


def _llm_extract(sample: str, model: str, base_url: str) -> dict:
    """LLM 提取摘要、標籤、實體（非串流，JSON 輸出）"""
    prompt = (
        '請分析以下文字，以JSON格式回傳（不要其他說明）：\n'
        '{"summary":"摘要100字內","tags":["tag1","tag2"],'
        '"entities":[{"name":"名稱","type":"PERSON|PLACE|ORG|CONCEPT","description":"簡述"}]}\n\n'
        f'文字：{sample[:2000]}'
    )
    with httpx.Client(timeout=180) as client:
        resp = client.post(
            f"{base_url}/api/generate",
            json={"model": model, "prompt": prompt, "stream": False},
        )
        resp.raise_for_status()
        raw = resp.json().get("response", "{}")
    m = re.search(r"\{.*\}", raw, re.DOTALL)
    if m:
        try:
            return json.loads(m.group())
        except json.JSONDecodeError:
            pass
    return {"summary": "", "tags": [], "entities": []}


def _llm_suggest_kb(
    content_sample: str,
    existing_kbs: list,
    model: str,
    base_url: str,
) -> dict:
    """
    根據文件內容建議最適合的 KB。
    回傳 {"kb_id": str | None, "kb_name": str}
    - existing_kbs 為空時：kb_id=None，kb_name=LLM 建議的新 KB 名稱
    - 有現有 KB 時：從中選最合適的，回傳其 id 和 name
    """
    if existing_kbs:
        kb_list_str = "\n".join(
            f"- id: {kb['id']}, 名稱: {kb['name']}, 描述: {kb.get('description') or '(無)'}"
            for kb in existing_kbs
        )
        prompt = (
            f'你是文件分類助理。請根據以下文件內容，'
            f'從現有知識庫中選出最適合的一個。\n\n'
            f'現有知識庫：\n{kb_list_str}\n\n'
            f'文件內容：\n{content_sample[:1000]}\n\n'
            f'請以 JSON 格式回傳（不要其他說明）：\n'
            f'{{"kb_id": "選中的KB ID", "kb_name": "選中的KB名稱"}}'
        )
    else:
        prompt = (
            f'你是文件分類助理。請根據以下文件內容，'
            f'建議一個適合的知識庫名稱（繁體中文，5-15字）。\n\n'
            f'文件內容：\n{content_sample[:1000]}\n\n'
            f'請以 JSON 格式回傳（不要其他說明）：\n'
            f'{{"kb_id": null, "kb_name": "建議的KB名稱"}}'
        )
    with httpx.Client(timeout=30) as client:
        resp = client.post(
            f"{base_url}/api/generate",
            json={"model": model, "prompt": prompt, "stream": False},
        )
        resp.raise_for_status()
        raw = resp.json().get("response", "{}")
    m = re.search(r"\{.*\}", raw, re.DOTALL)
    if m:
        try:
            result = json.loads(m.group())
            kb_id   = result.get("kb_id") or None
            kb_name = result.get("kb_name") or ""
            # 從 existing_kbs 驗證 kb_id 確實存在
            if kb_id and existing_kbs:
                valid_ids = {kb["id"] for kb in existing_kbs}
                if kb_id not in valid_ids:
                    kb_id = None
                    kb_name = existing_kbs[0]["name"]
            return {"kb_id": kb_id, "kb_name": kb_name}
        except json.JSONDecodeError:
            pass
    # 解析失敗：若有 KB 取第一個
    if existing_kbs:
        return {"kb_id": existing_kbs[0]["id"], "kb_name": existing_kbs[0]["name"]}
    return {"kb_id": None, "kb_name": ""}


def _llm_suggest_tags(
    content_sample: str,
    existing_tags: list,
    model: str,
    base_url: str,
) -> list:
    """
    根據文件內容從現有 tag 庫推薦標籤，或建議新標籤。
    回傳 [{'tag_id': str|None, 'tag_name': str, 'confidence': float}, ...]
    優先複用現有 tag（語意相近即直接用），只在現有 tag 無法描述時才建立新 tag。
    """
    if existing_tags:
        tag_names_str = "、".join(t["name"] for t in existing_tags)
        tag_list_str = "\n".join(
            f"- id: {t['id']}, 名稱: {t['name']}"
            + (f", 說明: {t['description']}" if t.get("description") else "")
            for t in existing_tags
        )
        prompt = (
            f"你是文件標籤助理。請根據以下文件內容，為文件選擇或建立標籤。\n\n"
            f"【可複用標籤清單】\n{tag_list_str}\n\n"
            f"【規則】\n"
            f"1. 優先從可複用清單選標籤，如果某個現有標籤與文件主題 70% 以上相關，直接使用它（tag_id 填入該標籤 id）\n"
            f"2. 只有在現有標籤完全無法描述文件某個重要主題時，才建議新標籤（tag_id 設為 null）\n"
            f"3. 新標籤名稱必須是清楚的主題詞（2-8 個字），不能是單字或過於籠統\n"
            f"4. 中文文章給中文標籤，英文文章給英文標籤\n"
            f"5. 不限數量，但每個標籤都要有明確意義\n\n"
            f"【文件內容】\n{content_sample[:1500]}\n\n"
            f"請以 JSON array 回傳（不要其他說明），依 confidence 降序：\n"
            f'[{{"tag_id": "現有的tag ID 或 null", "tag_name": "標籤名稱", "confidence": 0.9}}, ...]'
        )
    else:
        prompt = (
            f"你是文件標籤助理。請根據以下文件內容，建議適合的標籤。\n\n"
            f"【規則】\n"
            f"1. 標籤名稱必須是清楚的主題詞（2-8 個字），不能是單字或過於籠統\n"
            f"2. 中文文章給中文標籤，英文文章給英文標籤\n"
            f"3. 不限數量，但每個標籤都要有明確意義\n\n"
            f"【文件內容】\n{content_sample[:1500]}\n\n"
            f"請以 JSON array 回傳（不要其他說明），依 confidence 降序：\n"
            f'[{{"tag_id": null, "tag_name": "建議標籤名稱", "confidence": 0.9}}, ...]'
        )
    try:
        with httpx.Client(timeout=30) as client:
            resp = client.post(
                f"{base_url}/api/generate",
                json={"model": model, "prompt": prompt, "stream": False},
            )
            resp.raise_for_status()
            raw = resp.json().get("response", "[]")
        m = re.search(r"\[.*\]", raw, re.DOTALL)
        if not m:
            return []
        items = json.loads(m.group())
        if not isinstance(items, list):
            return []
        result = []
        valid_ids = {t["id"] for t in existing_tags}
        for item in items:
            tag_id   = item.get("tag_id") or None
            tag_name = (item.get("tag_name") or "").strip()
            confidence = float(item.get("confidence", 0.5))
            if not tag_name:
                continue
            # 驗證 tag_id 確實存在
            if tag_id and tag_id not in valid_ids:
                tag_id = None
            result.append({"tag_id": tag_id, "tag_name": tag_name, "confidence": round(confidence, 3)})
        return result
    except Exception:
        return []


def _find_similar_kbs(
    doc_id: str,
    threshold: float = 0.75,
) -> list:
    """
    用文件的 chunk embeddings 去 Qdrant 比對每個 KB 的平均相似度。
    回傳 [{"kb_id": str, "kb_name": str, "score": float}]（只回傳 score > threshold 的，降序）

    步驟：
    1. 從 PG 取出這份文件所有 chunks 的 vector_id
    2. 從 Qdrant 取出這些 vector
    3. 對每個現有 KB：
       - 用 filter(kb_id) scroll 取最多 50 個 chunk vectors 作採樣
       - 計算文件平均 vector 與 KB 平均 vector 的 cosine 相似度
    4. 回傳 score > threshold 的 KB，依 score 降序
    """
    import numpy as np
    from qdrant_client.models import Filter, FieldCondition, MatchValue

    s = _settings()

    # 1. 取文件 vector_ids
    with _pg_conn() as pg:
        with pg.cursor() as cur:
            cur.execute(
                "SELECT vector_id FROM chunks WHERE doc_id = %s AND vector_id IS NOT NULL",
                (doc_id,),
            )
            vector_ids = [row[0] for row in cur.fetchall()]

    if not vector_ids:
        return []

    # 2. 從 Qdrant 取文件 vectors
    qdrant = _qdrant()
    doc_points, _ = qdrant.scroll(
        collection_name=s.QDRANT_COLLECTION,
        scroll_filter=Filter(
            must=[FieldCondition(key="doc_id", match=MatchValue(value=doc_id))]
        ),
        with_vectors=True,
        limit=200,
    )
    if not doc_points:
        return []

    doc_vecs = np.array([p.vector for p in doc_points if p.vector is not None], dtype=np.float32)
    if doc_vecs.ndim == 1 or len(doc_vecs) == 0:
        return []
    doc_avg = doc_vecs.mean(axis=0)
    norm = np.linalg.norm(doc_avg)
    if norm > 0:
        doc_avg = doc_avg / norm

    # 3. 取現有所有 KB
    with _pg_conn() as pg:
        with pg.cursor() as cur:
            cur.execute("SELECT id::text, name FROM knowledge_bases ORDER BY created_at")
            kbs = [(row[0], row[1]) for row in cur.fetchall()]

    if not kbs:
        return []

    # 從 PG 取每個 KB 的 doc_id 清單（fallback 用）
    with _pg_conn() as pg:
        with pg.cursor() as cur:
            cur.execute(
                "SELECT kb_id::text, doc_id::text FROM document_knowledge_bases"
            )
            rows = cur.fetchall()
    kb_doc_ids: dict[str, list[str]] = {}
    for r_kb_id, r_doc_id in rows:
        kb_doc_ids.setdefault(r_kb_id, []).append(r_doc_id)

    results = []
    for kb_id, kb_name in kbs:
        # 優先用 kb_id payload filter
        kb_points, _ = qdrant.scroll(
            collection_name=s.QDRANT_COLLECTION,
            scroll_filter=Filter(
                must=[FieldCondition(key="kb_id", match=MatchValue(value=kb_id))]
            ),
            with_vectors=True,
            limit=50,
        )

        # Fallback：kb_id payload 缺失時，改用 PG doc_id 清單逐一 scroll
        if not kb_points:
            fallback_doc_ids = kb_doc_ids.get(kb_id, [])
            fallback_points = []
            for fb_doc_id in fallback_doc_ids[:10]:   # 最多取前 10 篇文件
                fb_pts, _ = qdrant.scroll(
                    collection_name=s.QDRANT_COLLECTION,
                    scroll_filter=Filter(
                        must=[FieldCondition(key="doc_id", match=MatchValue(value=fb_doc_id))]
                    ),
                    with_vectors=True,
                    limit=50,
                )
                fallback_points.extend(fb_pts)
                if len(fallback_points) >= 50:
                    break
            kb_points = fallback_points[:50]

        if not kb_points:
            continue

        kb_vecs = np.array([p.vector for p in kb_points if p.vector is not None], dtype=np.float32)
        if len(kb_vecs) == 0:
            continue
        kb_avg = kb_vecs.mean(axis=0)
        kb_norm = np.linalg.norm(kb_avg)
        if kb_norm > 0:
            kb_avg = kb_avg / kb_norm
        score = float(np.dot(doc_avg, kb_avg))
        if score >= threshold:
            results.append({"kb_id": kb_id, "kb_name": kb_name, "score": round(score, 4)})

    results.sort(key=lambda x: x["score"], reverse=True)
    return results


# ── 主任務 ────────────────────────────────────────────────────

@shared_task(
    bind=True,
    autoretry_for=(Exception,),
    max_retries=3,
    retry_backoff=True,
    retry_backoff_max=60,
    name="tasks.ingest_document",
)
def ingest_document(self, doc_id: str):
    """文件攝取主任務：解析 → LLM 分析 → 三庫寫入（Saga 保護）"""
    logger.info("[task:%s] Starting ingestion for doc_id=%s", self.request.id, doc_id)
    s = _settings()

    # 0. 確認文件尚未被刪除（用戶可能在 task 排隊期間取消）
    with _pg_conn() as pg:
        with pg.cursor() as cur:
            cur.execute("SELECT id FROM documents WHERE id = %s", (doc_id,))
            if cur.fetchone() is None:
                logger.info("[task:%s] doc_id=%s already deleted, skipping", self.request.id, doc_id)
                return

    # 1. PG 取文件資訊，狀態 → processing
    with _pg_conn() as pg:
        with pg.cursor() as cur:
            cur.execute(
                "UPDATE documents SET status='processing' WHERE id=%s "
                "RETURNING file_path, file_type, title, knowledge_base_id, source",
                (doc_id,),
            )
            row = cur.fetchone()
        pg.commit()

    if row is None:
        raise ValueError(f"Document {doc_id} not found in DB")
    file_path, file_type, title, knowledge_base_id, source_url = row

    # 2. MinIO 下載
    mc = _minio()
    response = mc.get_object(s.MINIO_BUCKET, file_path)
    data = response.read()
    response.close()
    response.release_conn()

    # 3. 解析
    parser = _PARSERS.get(file_type or "", _parse_text)
    pages  = parser(data)

    # 4. 分塊（chunk_size 優先取 KB 設定，fallback 全域）
    with _pg_conn() as _cfg_pg:
        _kb_cfg = _get_kb_config(_cfg_pg, knowledge_base_id)
        _dyn_chunk_size = _kb_cfg.get("chunk_size") or _get_chunk_size(_cfg_pg)
    if file_type == "xlsx":
        raw_chunks = _xlsx_chunks(pages)
    else:
        raw_chunks = []
        for text, page in pages:
            raw_chunks.extend(_sentence_window_chunks(text, page, chunk_size=_dyn_chunk_size))

    if not raw_chunks:
        with _pg_conn() as pg:
            with pg.cursor() as cur:
                cur.execute(
                    "UPDATE documents SET status='indexed', chunk_count=0 WHERE id=%s",
                    (doc_id,))
            pg.commit()
        logger.info("[task:%s] No chunks generated for doc_id=%s", self.request.id, doc_id)
        return

    # 5. 嵌入（依 KB 設定路由 ollama / 雲端）
    vectors   = _embed_dispatch([c["content"] for c in raw_chunks], _kb_cfg, doc_id=doc_id)
    chunk_ids = [str(uuid.uuid4()) for _ in raw_chunks]

    # 6-9. Saga 三庫寫入
    from services.saga import SagaLog
    saga = SagaLog("ingest_document", doc_id)
    saga.begin()

    qdrant_ids = [str(uuid.uuid4()) for _ in raw_chunks]

    try:
        # ── Qdrant ──────────────────────────────────────────
        qdrant = _qdrant()
        points = [
            PointStruct(
                id=qid,
                vector=vec,
                payload={
                    "doc_id":      doc_id,
                    "content":     c["content"],
                    "page_number": c["page_number"],
                    "title":       title,
                    "kb_id":       knowledge_base_id,
                    "source_url":  source_url,
                    "chunk_id":    cid,
                },
            )
            for qid, vec, c, cid in zip(qdrant_ids, vectors, raw_chunks, chunk_ids)
        ]
        qdrant.upsert(collection_name=s.QDRANT_COLLECTION, points=points)
        saga.record_step("qdrant")

        # ── LLM 分析（Ollama 不可用時降級為空，不阻斷索引）───────────
        sample = " ".join(c["content"] for c in raw_chunks[:10])[:3000]
        try:
            analysis = _llm_extract(sample, s.OLLAMA_LLM_MODEL, s.OLLAMA_BASE_URL)
        except Exception as llm_exc:
            logger.warning(
                "[task:%s] LLM entity extraction failed (degraded to empty): %s",
                self.request.id, llm_exc,
            )
            analysis = {"summary": "", "tags": [], "entities": []}

        # ── Neo4j ────────────────────────────────────────────
        neo_driver = _neo4j()
        with neo_driver.session(database="neo4j") as neo_sess:
            neo_sess.run(
                "MERGE (d:Document {id:$doc_id}) "
                "SET d.title=$title, d.tags=$tags, d.summary=$summary",
                doc_id=doc_id,
                title=title,
                tags=analysis.get("tags", []),
                summary=analysis.get("summary", ""),
            )
            for entity in analysis.get("entities", []):
                neo_sess.run(
                    """
                    MERGE (e:Entity {name: $name})
                    SET e.type=$etype, e.description=$desc
                    WITH e
                    MATCH (d:Document {id: $doc_id})
                    MERGE (d)-[:MENTIONS]->(e)
                    """,
                    name=entity.get("name", ""),
                    etype=entity.get("type", "CONCEPT"),
                    desc=entity.get("description", ""),
                    doc_id=doc_id,
                )
        neo_driver.close()
        saga.record_step("neo4j")

        # ── OntologyReviewQueue ──────────────────────────────
        try:
            with _pg_conn() as _orq_pg:
                with _orq_pg.cursor() as _orq_cur:
                    for entity in analysis.get("entities", []):
                        _orq_cur.execute(
                            """
                            INSERT INTO ontology_review_queue
                              (entity_name, entity_type, action, source_doc_id, status)
                            VALUES (%s, %s, 'create', %s::uuid, 'pending')
                            ON CONFLICT (entity_name, entity_type)
                            WHERE status = 'pending' DO NOTHING
                            """,
                            (entity.get("name", ""), entity.get("type", "CONCEPT"), doc_id),
                        )
                _orq_pg.commit()
        except Exception as rq_err:
            logger.warning("[task:%s] ReviewQueue 寫入失敗: %s", self.request.id, rq_err)

        # ── PG chunks ────────────────────────────────────────
        with _pg_conn() as pg:
            with pg.cursor() as cur:
                for i, (c, qid, cid) in enumerate(
                        zip(raw_chunks, qdrant_ids, chunk_ids)):
                    cur.execute(
                        """
                        INSERT INTO chunks
                          (id, doc_id, content, chunk_index, vector_id, window_context, page_number)
                        VALUES (%s, %s, %s, %s, %s, %s, %s)
                        """,
                        (cid, doc_id, c["content"], i, qid,
                         c["window_context"], c["page_number"]),
                    )
                cur.execute(
                    "UPDATE documents SET status='indexed', chunk_count=%s WHERE id=%s",
                    (len(raw_chunks), doc_id),
                )
            pg.commit()
        saga.record_step("postgres")

        saga.commit()
        logger.info(
            "[task:%s] Indexed doc_id=%s: %d chunks",
            self.request.id, doc_id, len(raw_chunks)
        )

        # ── 自動更新跨文件關聯（優點2：零維護）──────────────────────
        try:
            neo_driver2 = _neo4j()
            with neo_driver2.session(database="neo4j") as neo_sess:
                # 與新文件共享相同 Entity 的所有文件 → 建立 RELATED_TO
                neo_sess.run(
                    """
                    MATCH (d1:Document {id: $doc_id})-[:MENTIONS]->(e:Entity)<-[:MENTIONS]-(d2:Document)
                    WHERE d1 <> d2
                    MERGE (d1)-[r:RELATED_TO]->(d2)
                    SET r.updated_at = datetime()
                    """,
                    doc_id=doc_id,
                )
            neo_driver2.close()
            logger.info("[task:%s] Updated RELATED_TO for doc_id=%s", self.request.id, doc_id)
        except Exception as link_exc:
            logger.warning(
                "[task:%s] RELATED_TO update skipped for doc_id=%s: %s",
                self.request.id, doc_id, link_exc,
            )

        # ── LLM 自動 KB 分類（只在未指定 KB 時執行）────────────────────
        if knowledge_base_id is None:
            try:
                similar_kbs = _find_similar_kbs(doc_id, threshold=s.KB_SIMILARITY_THRESHOLD)

                if similar_kbs:
                    # 有相似 KB → 寫入 document_knowledge_bases
                    with _pg_conn() as _dkb_pg:
                        with _dkb_pg.cursor() as _dkb_cur:
                            for kb_hit in similar_kbs:
                                _dkb_cur.execute(
                                    """
                                    INSERT INTO document_knowledge_bases (doc_id, kb_id, score, source)
                                    VALUES (%s::uuid, %s::uuid, %s, 'auto')
                                    ON CONFLICT (doc_id, kb_id) DO UPDATE SET score = EXCLUDED.score
                                    """,
                                    (doc_id, kb_hit["kb_id"], kb_hit["score"]),
                                )
                            # 同步更新 documents.knowledge_base_id 為 score 最高的（向下相容）
                            _dkb_cur.execute(
                                "UPDATE documents SET knowledge_base_id = %s::uuid WHERE id = %s::uuid",
                                (similar_kbs[0]["kb_id"], doc_id),
                            )
                        _dkb_pg.commit()
                    logger.info(
                        "[task:%s] KB auto-classified doc=%s → %d KB(s), top='%s' score=%.4f",
                        self.request.id, doc_id, len(similar_kbs),
                        similar_kbs[0]["kb_name"], similar_kbs[0]["score"],
                    )
                else:
                    # 無相似 KB → LLM 建議名稱 → 直接建立新 KB → 歸入文件 → 前端待審核
                    _tag_names_for_kb = []
                    with _pg_conn() as _tn_pg:
                        with _tn_pg.cursor() as _tn_cur:
                            _tn_cur.execute(
                                "SELECT t.name FROM document_tags dt "
                                "JOIN tags t ON t.id = dt.tag_id "
                                "WHERE dt.doc_id = %s",
                                (doc_id,),
                            )
                            _tag_names_for_kb = [r[0] for r in _tn_cur.fetchall()]

                    # 取文件 created_by（新 KB 的擁有者）
                    _doc_created_by = None
                    with _pg_conn() as _dcb_pg:
                        with _dcb_pg.cursor() as _dcb_cur:
                            _dcb_cur.execute(
                                "SELECT created_by FROM documents WHERE id = %s",
                                (doc_id,),
                            )
                            _dcb_row = _dcb_cur.fetchone()
                            if _dcb_row:
                                _doc_created_by = _dcb_row[0]

                    # LLM 建議 KB 名稱（傳入 tag 清單 + 文件摘要）
                    _doc_summary = (analysis or {}).get("summary", "")
                    _sug_kb_name = ""
                    _kb_prompt = (
                        "你是文件分類助理。請根據以下資訊，用 2-10 個字為知識庫命名"
                        "（繁體中文，只回傳名稱，不要任何解釋）：\n"
                        f"標籤：{', '.join(_tag_names_for_kb) if _tag_names_for_kb else '（無）'}\n"
                        f"文件摘要：{_doc_summary[:200] if _doc_summary else '（無）'}"
                    )
                    try:
                        with httpx.Client(timeout=20) as _kc:
                            _kr = _kc.post(
                                f"{s.OLLAMA_BASE_URL}/api/generate",
                                json={"model": s.OLLAMA_LLM_MODEL, "prompt": _kb_prompt, "stream": False},
                            )
                            _kr.raise_for_status()
                            _sug_kb_name = _kr.json().get("response", "").strip().strip('"').strip()
                    except Exception:
                        pass

                    # Fallback：LLM 失敗時用文件標題前 20 字
                    if not _sug_kb_name:
                        _sug_kb_name = (title or "未命名文件")[:20]

                    # 建立新 KB，歸入文件，更新 documents 三個欄位
                    _new_kb_id = str(uuid.uuid4())
                    with _pg_conn() as _nkb_pg:
                        with _nkb_pg.cursor() as _nkb_cur:
                            _nkb_cur.execute(
                                """
                                INSERT INTO knowledge_bases (id, name, description, created_by)
                                VALUES (%s::uuid, %s, 'AI 自動建立', %s::uuid)
                                """,
                                (_new_kb_id, _sug_kb_name, _doc_created_by),
                            )
                            _nkb_cur.execute(
                                """
                                INSERT INTO document_knowledge_bases (doc_id, kb_id, score, source)
                                VALUES (%s::uuid, %s::uuid, 1.0, 'auto')
                                ON CONFLICT (doc_id, kb_id) DO NOTHING
                                """,
                                (doc_id, _new_kb_id),
                            )
                            _nkb_cur.execute(
                                """
                                UPDATE documents
                                SET knowledge_base_id = %s::uuid,
                                    suggested_kb_id   = %s::uuid,
                                    suggested_kb_name = %s
                                WHERE id = %s
                                """,
                                (_new_kb_id, _new_kb_id, _sug_kb_name, doc_id),
                            )
                        _nkb_pg.commit()
                    logger.info(
                        "[task:%s] No similar KB found for doc=%s; created new KB id=%s name='%s'",
                        self.request.id, doc_id, _new_kb_id, _sug_kb_name,
                    )
            except Exception as _kb_exc:
                logger.warning(
                    "[task:%s] KB classification failed for doc=%s, skipping: %s",
                    self.request.id, doc_id, _kb_exc,
                )

        # ── LLM 自動貼標籤（無論是否已有 KB 都執行）───────────────────
        try:
            with _pg_conn() as _tag_pg:
                with _tag_pg.cursor() as _tag_cur:
                    _tag_cur.execute(
                        "SELECT id::text, name, description FROM tags ORDER BY created_at"
                    )
                    existing_tags = [
                        {"id": r[0], "name": r[1], "description": r[2]}
                        for r in _tag_cur.fetchall()
                    ]

            _tag_content_sample = " ".join(c["content"] for c in raw_chunks[:5])[:1000]
            tag_suggestions = _llm_suggest_tags(
                _tag_content_sample, existing_tags, s.OLLAMA_LLM_MODEL, s.OLLAMA_BASE_URL
            )

            # 對 tag_id=None 的建議，自動建立新 tag
            import re as _re
            for _sug in tag_suggestions:
                if _sug["tag_id"] is None and _sug["tag_name"]:
                    _slug = _sug["tag_name"].strip().lower()
                    _slug = _re.sub(r"[\s_]+", "-", _slug)
                    _slug = _re.sub(r"[^\w\u4e00-\u9fff-]", "", _slug)
                    _new_tag_id = str(uuid.uuid4())
                    with _pg_conn() as _ntpg:
                        with _ntpg.cursor() as _ntcur:
                            _ntcur.execute(
                                """
                                INSERT INTO tags (id, name, slug, color)
                                VALUES (%s, %s, %s, '#409eff')
                                ON CONFLICT (slug) DO NOTHING
                                """,
                                (_new_tag_id, _sug["tag_name"].strip(), _slug),
                            )
                            _ntcur.execute(
                                "SELECT id::text FROM tags WHERE slug = %s", (_slug,)
                            )
                            _row = _ntcur.fetchone()
                        _ntpg.commit()
                    if _row:
                        _sug["tag_id"] = _row[0]

            with _pg_conn() as _tsave_pg:
                with _tsave_pg.cursor() as _tsave_cur:
                    _tsave_cur.execute(
                        "UPDATE documents SET suggested_tags = %s::jsonb WHERE id = %s",
                        (json.dumps(tag_suggestions), doc_id),
                    )
                _tsave_pg.commit()

            logger.info(
                "[task:%s] Tag suggestions: doc=%s → %d tags",
                self.request.id, doc_id, len(tag_suggestions),
            )
        except Exception as _tag_exc:
            logger.warning(
                "[task:%s] Tag suggestion failed for doc=%s, skipping: %s",
                self.request.id, doc_id, _tag_exc, exc_info=True,
            )

    except Exception as exc:
        saga.mark_compensated(error=str(exc))
        with _pg_conn() as pg:
            with pg.cursor() as cur:
                cur.execute(
                    "UPDATE documents SET status='failed', error_message=%s WHERE id=%s",
                    (str(exc)[:500], doc_id),
                )
            pg.commit()
        logger.error(
            "[task:%s] Ingestion failed for doc_id=%s: %s",
            self.request.id, doc_id, exc, exc_info=True
        )
        raise
