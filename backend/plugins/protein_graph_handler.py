"""
Protein Graph Handler
支援動作:
  import_genes  — 解析 GeneCards URL xlsx，寫入 proteins 表
  import_scores — 解析連結評分 xlsx（4 sheet），寫入 protein_interactions 表
  get_graph     — 取得指定 network 的節點 + 邊 (JSON)
  get_networks  — 列出已匯入的 network 名稱
  get_top       — 取前 N 高分邊
"""
import io
import logging
import math
import re
import urllib.parse
from typing import Any

logger = logging.getLogger(__name__)

# ──────────────────────────────────────────────────────────────
# 工具函式
# ──────────────────────────────────────────────────────────────

def _extract_symbol(url: str) -> str | None:
    """從 GeneCards URL 解析基因符號，e.g. ?gene=AKT1 → AKT1"""
    try:
        qs = urllib.parse.urlparse(url.strip()).query
        params = urllib.parse.parse_qs(qs)
        return params.get("gene", [None])[0]
    except Exception:
        return None


def _parse_genes_xlsx(data: bytes) -> list[dict]:
    """解析 genes list xlsx，回傳 [{'symbol': 'AKT1', 'genecards_url': '...'}]"""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl not installed")

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    results = []
    for row in rows:
        for cell in row:
            if cell is None:
                continue
            val = str(cell).strip()
            if "genecards.org" not in val:
                continue
            sym = _extract_symbol(val)
            if sym:
                results.append({"symbol": sym.upper(), "genecards_url": val})
    return results


def _parse_scores_xlsx(data: bytes) -> list[dict]:
    """
    解析 scores xlsx（上三角矩陣，NaN=無資料）
    回傳 [{'protein_a': 'A', 'protein_b': 'B', 'score': 0.9, 'network': 'USP7'}]
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl not installed")

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    interactions = []

    for sheet_name in wb.sheetnames:
        # network 名稱 = sheet name 去掉 " data"
        network = re.sub(r"\s+data$", "", sheet_name, flags=re.IGNORECASE).strip()
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        # 第一列為 header（col 0 是空/Unnamed, col 1+ 是基因符號）
        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        col_symbols = header[1:]  # 去掉第一欄

        for row in rows[1:]:
            row_symbol = str(row[0]).strip() if row[0] is not None else ""
            if not row_symbol:
                continue
            for ci, val in enumerate(row[1:]):
                if val is None:
                    continue
                try:
                    score = float(val)
                except (TypeError, ValueError):
                    continue
                if math.isnan(score) or score <= 0:
                    continue
                col_sym = col_symbols[ci] if ci < len(col_symbols) else ""
                if not col_sym or col_sym == row_symbol:
                    continue
                # 上三角：row_symbol < col_sym 字典序，確保唯一
                a, b = sorted([row_symbol.upper(), col_sym.upper()])
                interactions.append({
                    "protein_a": a,
                    "protein_b": b,
                    "score":     round(score, 4),
                    "network":   network,
                })

    wb.close()
    # 去重（同 a/b/network 保留最高分）
    seen: dict[tuple, float] = {}
    for item in interactions:
        key = (item["protein_a"], item["protein_b"], item["network"])
        if seen.get(key, -1) < item["score"]:
            seen[key] = item["score"]

    return [
        {"protein_a": k[0], "protein_b": k[1], "network": k[2], "score": v}
        for k, v in seen.items()
    ]


# ──────────────────────────────────────────────────────────────
# DB helpers (raw asyncpg via SQLAlchemy text)
# ──────────────────────────────────────────────────────────────

async def _db_upsert_proteins(proteins: list[dict], db) -> int:
    from sqlalchemy import text
    count = 0
    for p in proteins:
        await db.execute(
            text("""
                INSERT INTO proteins (symbol, genecards_url)
                VALUES (:symbol, :url)
                ON CONFLICT (symbol) DO UPDATE SET genecards_url = EXCLUDED.genecards_url
            """),
            {"symbol": p["symbol"], "url": p["genecards_url"]},
        )
        count += 1
    await db.commit()
    return count


async def _db_upsert_interactions(interactions: list[dict], db) -> int:
    from sqlalchemy import text
    count = 0
    for i in interactions:
        await db.execute(
            text("""
                INSERT INTO protein_interactions (protein_a, protein_b, score, network)
                VALUES (:a, :b, :score, :network)
                ON CONFLICT (protein_a, protein_b, network)
                DO UPDATE SET score = EXCLUDED.score
            """),
            {"a": i["protein_a"], "b": i["protein_b"], "score": i["score"], "network": i["network"]},
        )
        count += 1
    await db.commit()
    return count


async def _get_networks(db) -> list[str]:
    from sqlalchemy import text
    result = await db.execute(
        text("SELECT DISTINCT network FROM protein_interactions ORDER BY network")
    )
    return [r[0] for r in result.fetchall()]


async def _get_graph_data(network: str, min_score: float, db) -> dict:
    from sqlalchemy import text

    # 取邊
    result = await db.execute(
        text("""
            SELECT protein_a, protein_b, score
            FROM protein_interactions
            WHERE network = :network AND score >= :min_score
            ORDER BY score DESC
            LIMIT 500
        """),
        {"network": network, "min_score": min_score},
    )
    edges_raw = result.fetchall()

    # 蒐集節點
    node_set: dict[str, dict] = {}
    edges = []
    for a, b, score in edges_raw:
        for sym in [a, b]:
            if sym not in node_set:
                node_set[sym] = {"id": sym, "name": sym, "val": 1}
            else:
                node_set[sym]["val"] += 1
        edges.append({"source": a, "target": b, "value": score})

    # 查詢 genecards_url
    if node_set:
        syms = list(node_set.keys())
        placeholders = ", ".join(f":s{i}" for i in range(len(syms)))
        params = {f"s{i}": s for i, s in enumerate(syms)}
        res2 = await db.execute(
            text(f"SELECT symbol, genecards_url FROM proteins WHERE symbol IN ({placeholders})"),
            params,
        )
        for sym, url in res2.fetchall():
            if sym in node_set:
                node_set[sym]["url"] = url

    return {
        "nodes": list(node_set.values()),
        "links": edges,
        "network": network,
    }


async def _get_top_interactions(network: str, limit: int, db) -> list[dict]:
    from sqlalchemy import text
    result = await db.execute(
        text("""
            SELECT protein_a, protein_b, score, network
            FROM protein_interactions
            WHERE network = :network
            ORDER BY score DESC
            LIMIT :limit
        """),
        {"network": network, "limit": limit},
    )
    return [
        {"protein_a": r[0], "protein_b": r[1], "score": r[2], "network": r[3]}
        for r in result.fetchall()
    ]


# ──────────────────────────────────────────────────────────────
# 主入口
# ──────────────────────────────────────────────────────────────

async def run(action: str, params: dict, config: dict) -> dict:
    """Plugin handler 主入口（由 registry.dispatch 呼叫）"""
    from database import AsyncSessionLocal

    if action == "get_networks":
        async with AsyncSessionLocal() as db:
            networks = await _get_networks(db)
        return {"success": True, "networks": networks}

    if action == "get_graph":
        network   = params.get("network", "USP7")
        min_score = float(params.get("min_score", 0.5))
        async with AsyncSessionLocal() as db:
            data = await _get_graph_data(network, min_score, db)
        return {"success": True, **data}

    if action == "get_top":
        network = params.get("network", "USP7")
        limit   = int(params.get("limit", 20))
        async with AsyncSessionLocal() as db:
            rows = await _get_top_interactions(network, limit, db)
        return {"success": True, "interactions": rows}

    return {"success": False, "error": f"未知 action: {action}"}
