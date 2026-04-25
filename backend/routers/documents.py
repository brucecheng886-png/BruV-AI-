"""
文件 Router — 上傳、狀態查詢、列表、刪除
"""
import logging
import uuid
from typing import Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from pydantic import BaseModel
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from sqlalchemy.dialects.postgresql import insert as pg_insert

from auth import CurrentUser
from database import get_db
from models import Document, DocumentKnowledgeBase
from services.storage import upload_file
from tasks.document_tasks import ingest_document

logger = logging.getLogger(__name__)
router = APIRouter()

ALLOWED_TYPES = {
    "application/pdf": "pdf",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": "docx",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "xlsx",
    "text/plain": "txt",
    "text/markdown": "md",
    "text/html": "html",
    "text/csv": "csv",
}
MAX_FILE_SIZE = 50 * 1024 * 1024  # 50 MB


class TagSuggestionAction(BaseModel):
    action: str          # 'confirm' | 'reject'
    tag_ids: list[str] = []   # confirm 時選取的 tag_ids


class DocumentOut(BaseModel):
    doc_id: str
    title: str
    status: str
    file_type: str | None
    chunk_count: int
    knowledge_base_id: str | None = None
    knowledge_base_name: str | None = None
    suggested_kb_id: Optional[str] = None
    suggested_kb_name: Optional[str] = None
    suggested_tags: list[dict] = []   # [{"tag_id": str, "tag_name": str, "confidence": float}]
    tags: list[str] = []
    kb_list: list[dict] = []          # [{"kb_id": str, "kb_name": str, "score": float|None, "source": str}]
    cover_image_url: Optional[str] = None
    source_url: Optional[str] = None
    created_at: str

    class Config:
        from_attributes = True


class DocumentStatusOut(BaseModel):
    doc_id: str
    status: str
    chunk_count: int
    error_message: str | None


@router.post("/upload", status_code=status.HTTP_202_ACCEPTED)
async def upload_document(
    file: UploadFile = File(...),
    knowledge_base_id: Optional[str] = Form(None),
    current_user: CurrentUser = None,
    db: AsyncSession = Depends(get_db),
):
    """上傳文件並觸發非同步攝取任務"""
    # 1. 驗證 content type
    content_type = file.content_type or ""
    file_type = ALLOWED_TYPES.get(content_type)
    if file_type is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"不支援的檔案類型：{content_type}。支援：{list(ALLOWED_TYPES.keys())}",
        )

    # 2. 讀取並驗證大小
    data = await file.read()
    if len(data) > MAX_FILE_SIZE:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"檔案過大（{len(data) // 1024 // 1024} MB），上限 50 MB",
        )

    # 3. 建立 Document 記錄
    doc_id = str(uuid.uuid4())
    title = file.filename or doc_id
    object_name = f"documents/{doc_id}/{file.filename}"

    # 4. 上傳至 MinIO
    upload_file(object_name, data, content_type)

    # 5. 寫入 PG
    doc = Document(
        id=doc_id,
        title=title,
        source=file.filename,
        file_path=object_name,
        file_type=file_type,
        status="pending",
        created_by=current_user.id if current_user else None,
        knowledge_base_id=knowledge_base_id or None,
    )
    db.add(doc)
    await db.commit()

    # 6. 同步 M2M 表
    if knowledge_base_id:
        await db.execute(
            pg_insert(DocumentKnowledgeBase).values(
                doc_id=doc_id, kb_id=knowledge_base_id, source="manual"
            ).on_conflict_do_nothing(index_elements=["doc_id", "kb_id"])
        )
        await db.commit()

    # 7. 觸發 Celery 攝取任務
    ingest_document.delay(doc_id)

    logger.info("Document uploaded: %s (%s)", doc_id, file.filename)
    return {"doc_id": doc_id, "title": title, "status": "pending"}


@router.post("/import-excel", status_code=status.HTTP_202_ACCEPTED)
async def import_excel(
    file: UploadFile = File(...),
    knowledge_base_id: Optional[str] = Form(None),
    current_user: CurrentUser = None,
    db: AsyncSession = Depends(get_db),
):
    """
    批次匯入 Excel 中的連結並觸發爬取任務。
    Excel 必要欄位：srl / title / link（description / Page 為選用）
    回傳：{total, queued, skipped, skipped_items}
    """
    from io import BytesIO
    import re as _re
    from openpyxl import load_workbook

    content_type = file.content_type or ""
    if "spreadsheetml" not in content_type and not file.filename.endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="僅支援 .xlsx 檔案",
        )

    data = await file.read()
    if len(data) > MAX_FILE_SIZE:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"檔案過大（{len(data) // 1024 // 1024} MB），上限 50 MB",
        )

    # 解析 Excel
    try:
        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Excel 解析失敗：{exc}")

    if not rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel 為空")

    # 解析欄位對應（不區分大小寫）
    header = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    def _col(name: str) -> int | None:
        try:
            return header.index(name)
        except ValueError:
            return None

    col_srl   = _col("srl")
    col_title = _col("title")
    col_link  = _col("link")
    col_img   = _col("imgsrc")

    if col_link is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel 缺少 'link' 欄位",
        )

    # 不支援的 URL 關鍵字及對應原因
    _BLOCKED: list[tuple[str, str]] = [
        ("facebook.com",      "facebook.com 不支援"),
        ("youtube.com",       "youtube.com 不支援"),
        ("youtu.be",          "youtube.com 不支援"),
        ("google.com/drive",  "Google Drive 不支援"),
        ("drive.google.com",  "Google Drive 不支援"),
        ("blob:",             "blob: URL 不支援"),
    ]

    queued_ids: list[str] = []
    skipped_items: list[dict] = []

    for row in rows[1:]:
        if all(v is None for v in row):
            continue

        srl   = str(row[col_srl]).strip()   if col_srl   is not None and row[col_srl]   is not None else ""
        title = str(row[col_title]).strip() if col_title is not None and row[col_title] is not None else ""
        link  = str(row[col_link]).strip()  if row[col_link] is not None else ""
        img_src = str(row[col_img]).strip() if col_img is not None and row[col_img] is not None else None
        if img_src == "" or img_src == "None":
            img_src = None

        if not link:
            skipped_items.append({"srl": srl, "title": title, "reason": "連結為空"})
            continue

        # URL 格式驗證
        if not link.startswith(("http://", "https://")):
            skipped_items.append({"srl": srl, "title": title, "reason": "無效 URL（需 http/https）"})
            continue

        # 封鎖清單
        blocked_reason = next(
            (reason for pattern, reason in _BLOCKED if pattern in link.lower()),
            None,
        )
        if blocked_reason:
            skipped_items.append({"srl": srl, "title": title, "reason": blocked_reason})
            continue

        # 建立 Document 記錄並觸發爬取
        doc_id = str(uuid.uuid4())
        row_title = title or link[:200]
        doc = Document(
            id=doc_id,
            title=row_title,
            source=link,
            file_type="html",
            status="pending",
            created_by=current_user.id if current_user else None,
            knowledge_base_id=knowledge_base_id or None,
            cover_image_url=img_src,
        )
        db.add(doc)
        queued_ids.append(doc_id)

    await db.commit()

    # 同步 M2M 表（批次）
    if knowledge_base_id and queued_ids:
        for qid in queued_ids:
            await db.execute(
                pg_insert(DocumentKnowledgeBase).values(
                    doc_id=qid, kb_id=knowledge_base_id, source="manual"
                ).on_conflict_do_nothing(index_elements=["doc_id", "kb_id"])
            )
        await db.commit()

    # 重建 queued doc_id → link 對應後觸發 Celery 任務
    db_docs = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        link  = str(row[col_link]).strip() if row[col_link] is not None else ""
        title = str(row[col_title]).strip() if col_title is not None and row[col_title] is not None else ""
        if not link or not link.startswith(("http://", "https://")):
            continue
        if any(p in link.lower() for p, _ in _BLOCKED):
            continue
        db_docs.append((link, title or link[:200]))

    from tasks.crawl_tasks import crawl_document as _crawl
    for doc_id, (link, _title) in zip(queued_ids, db_docs):
        _crawl.delay(doc_id, link)

    logger.info(
        "import-excel: total=%d queued=%d skipped=%d user=%s",
        len(rows) - 1, len(queued_ids), len(skipped_items),
        current_user.id if current_user else "anon",
    )
    return {
        "total":         len(rows) - 1,
        "queued":        len(queued_ids),
        "skipped":       len(skipped_items),
        "skipped_items": skipped_items,
    }


@router.get("/count")
async def count_documents(
    status_filter: Optional[str] = Query(default=None, alias="status"),
    kb_id: Optional[str] = Query(default=None),
    db: AsyncSession = Depends(get_db),
):
    """回傳符合條件的文件總筆數（預設排除已刪除）"""
    q = select(func.count()).select_from(Document).where(Document.deleted_at.is_(None))
    if status_filter:
        q = q.where(Document.status == status_filter)
    if kb_id == "__none__":
        q = q.where(Document.knowledge_base_id.is_(None))
    elif kb_id:
        q = q.where(Document.knowledge_base_id == kb_id)
    result = await db.execute(q)
    return {"total": result.scalar() or 0}


@router.get("/{doc_id}/status", response_model=DocumentStatusOut)
async def get_document_status(doc_id: str, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Document).where(Document.id == doc_id))
    doc = result.scalar_one_or_none()
    if doc is None:
        raise HTTPException(status_code=404, detail="文件不存在")
    return DocumentStatusOut(
        doc_id=doc.id,
        status=doc.status,
        chunk_count=doc.chunk_count,
        error_message=doc.error_message,
    )


@router.get("/trash", response_model=list[DocumentOut])
async def list_trash(
    limit: int = Query(default=50, le=200),
    offset: int = Query(default=0, ge=0),
    db: AsyncSession = Depends(get_db),
):
    """列出垃圾桶（已軟刪除的文件）"""
    from models import KnowledgeBase, DocumentTag, Tag

    q = (
        select(Document)
        .where(Document.deleted_at.is_not(None))
        .order_by(Document.deleted_at.desc())
        .limit(limit)
        .offset(offset)
    )
    result = await db.execute(q)
    docs = result.scalars().all()
    if not docs:
        return []

    doc_ids = [d.id for d in docs]
    kb_ids = {d.knowledge_base_id for d in docs if d.knowledge_base_id}
    kb_map: dict[str, str] = {}
    if kb_ids:
        kb_result = await db.execute(select(KnowledgeBase).where(KnowledgeBase.id.in_(kb_ids)))
        kb_map = {kb.id: kb.name for kb in kb_result.scalars().all()}

    return [
        DocumentOut(
            doc_id=d.id,
            title=d.title,
            status=d.status,
            file_type=d.file_type,
            chunk_count=d.chunk_count,
            knowledge_base_id=d.knowledge_base_id,
            knowledge_base_name=kb_map.get(d.knowledge_base_id) if d.knowledge_base_id else None,
            suggested_kb_id=d.suggested_kb_id,
            suggested_kb_name=d.suggested_kb_name,
            suggested_tags=d.suggested_tags or [],
            tags=[],
            kb_list=[],
            cover_image_url=d.cover_image_url,
            source_url=d.source or None,
            created_at=d.created_at.isoformat(),
        )
        for d in docs
    ]


@router.get("/", response_model=list[DocumentOut])
async def list_documents(
    limit: int = Query(default=20, le=100),
    offset: int = Query(default=0, ge=0),
    status_filter: Optional[str] = Query(default=None, alias="status"),
    kb_id: Optional[str] = Query(default=None),
    tag_id: Optional[str] = Query(default=None),
    include_deleted: bool = Query(default=False),
    db: AsyncSession = Depends(get_db),
):
    from models import KnowledgeBase, DocumentTag, Tag

    q = select(Document).order_by(Document.created_at.desc()).limit(limit).offset(offset)
    if not include_deleted:
        q = q.where(Document.deleted_at.is_(None))
    if status_filter:
        q = q.where(Document.status == status_filter)
    if kb_id == "__none__":
        q = q.where(Document.knowledge_base_id.is_(None))
    elif kb_id:
        q = q.where(Document.knowledge_base_id == kb_id)
    if tag_id:
        q = q.where(
            Document.id.in_(
                select(DocumentTag.doc_id).where(DocumentTag.tag_id == tag_id)
            )
        )
    result = await db.execute(q)
    docs = result.scalars().all()

    if not docs:
        return []

    doc_ids = [d.id for d in docs]

    # batch load KB names（向下相容用，從 knowledge_base_id）
    kb_ids = {d.knowledge_base_id for d in docs if d.knowledge_base_id}
    kb_map: dict[str, str] = {}
    if kb_ids:
        kb_result = await db.execute(select(KnowledgeBase).where(KnowledgeBase.id.in_(kb_ids)))
        kb_map = {kb.id: kb.name for kb in kb_result.scalars().all()}

    # batch load kb_list（從 document_knowledge_bases 多對多表）
    from models import DocumentKnowledgeBase
    from sqlalchemy import and_
    dkb_result = await db.execute(
        select(DocumentKnowledgeBase, KnowledgeBase.name)
        .join(KnowledgeBase, KnowledgeBase.id == DocumentKnowledgeBase.kb_id)
        .where(DocumentKnowledgeBase.doc_id.in_(doc_ids))
        .order_by(DocumentKnowledgeBase.doc_id, DocumentKnowledgeBase.score.desc().nullslast())
    )
    kb_list_map: dict[str, list[dict]] = {}
    for dkb, kb_name in dkb_result:
        kb_list_map.setdefault(dkb.doc_id, []).append({
            "kb_id": dkb.kb_id,
            "kb_name": kb_name,
            "score": dkb.score,
            "source": dkb.source,
        })

    # batch load tags（避免 N+1）
    tags_result = await db.execute(
        select(DocumentTag.doc_id, Tag.name)
        .join(Tag, Tag.id == DocumentTag.tag_id)
        .where(DocumentTag.doc_id.in_(doc_ids))
    )
    tags_map: dict[str, list[str]] = {}
    for row in tags_result:
        tags_map.setdefault(row.doc_id, []).append(row.name)

    return [
        DocumentOut(
            doc_id=d.id,
            title=d.title,
            status=d.status,
            file_type=d.file_type,
            chunk_count=d.chunk_count,
            knowledge_base_id=d.knowledge_base_id,
            knowledge_base_name=kb_map.get(d.knowledge_base_id) if d.knowledge_base_id else None,
            suggested_kb_id=d.suggested_kb_id,
            suggested_kb_name=d.suggested_kb_name,
            suggested_tags=d.suggested_tags or [],
            tags=tags_map.get(d.id, []),
            kb_list=kb_list_map.get(d.id, []),
            cover_image_url=d.cover_image_url,
            source_url=d.source or None,
            created_at=d.created_at.isoformat(),
        )
        for d in docs
    ]


@router.patch("/{doc_id}/knowledge-base")
async def confirm_kb_suggestion(
    doc_id: str,
    body: dict,
    current_user: CurrentUser = None,
    db: AsyncSession = Depends(get_db),
):
    """
    確認或拒絕 LLM 建議的 KB。
    body: {"action": "confirm"} → knowledge_base_id = suggested_kb_id，清空建議欄位
    body: {"action": "reject"}  → 只清空建議欄位，knowledge_base_id 不動
    """
    result = await db.execute(select(Document).where(Document.id == doc_id))
    doc = result.scalar_one_or_none()
    if doc is None:
        raise HTTPException(status_code=404, detail="文件不存在")

    action = body.get("action")
    if action not in ("confirm", "reject"):
        raise HTTPException(status_code=400, detail="action 必須為 confirm 或 reject")

    if action == "confirm":
        if doc.suggested_kb_id is None:
            raise HTTPException(status_code=400, detail="沒有待確認的建議")
        doc.knowledge_base_id = doc.suggested_kb_id
        # 同步 M2M 表
        await db.execute(
            pg_insert(DocumentKnowledgeBase).values(
                doc_id=doc_id, kb_id=doc.suggested_kb_id, source="manual"
            ).on_conflict_do_nothing(index_elements=["doc_id", "kb_id"])
        )

    doc.suggested_kb_id   = None
    doc.suggested_kb_name = None

    await db.commit()
    return {
        "doc_id": doc_id,
        "action": action,
        "knowledge_base_id": str(doc.knowledge_base_id) if doc.knowledge_base_id else None,
    }


@router.patch("/{doc_id}/knowledge-bases")
async def update_document_knowledge_bases(
    doc_id: str,
    body: dict,
    current_user: CurrentUser = None,
    db: AsyncSession = Depends(get_db),
):
    """
    確認或拒絕多對多 KB 分配。
    body: {"action": "confirm", "kb_ids": ["id1", "id2"]}
      → 寫入 document_knowledge_bases（source='manual'），清空 suggested_kb_name/id
    body: {"action": "reject"}
      → 只清空 suggested_kb_name/id
    """
    from models import DocumentKnowledgeBase, KnowledgeBase

    result = await db.execute(select(Document).where(Document.id == doc_id))
    doc = result.scalar_one_or_none()
    if doc is None:
        raise HTTPException(status_code=404, detail="文件不存在")

    action = body.get("action")
    if action not in ("confirm", "reject"):
        raise HTTPException(status_code=400, detail="action 必須為 confirm 或 reject")

    if action == "confirm":
        kb_ids: list[str] = body.get("kb_ids") or []
        if not kb_ids:
            raise HTTPException(status_code=400, detail="confirm 時必須提供 kb_ids")

        # 驗證所有 kb_id 確實存在
        kb_result = await db.execute(
            select(KnowledgeBase).where(KnowledgeBase.id.in_(kb_ids))
        )
        valid_kbs = {kb.id: kb for kb in kb_result.scalars().all()}

        for kid in kb_ids:
            if kid not in valid_kbs:
                continue
            existing = await db.execute(
                select(DocumentKnowledgeBase).where(
                    DocumentKnowledgeBase.doc_id == doc_id,
                    DocumentKnowledgeBase.kb_id == kid,
                )
            )
            if existing.scalar_one_or_none() is None:
                dkb = DocumentKnowledgeBase(
                    doc_id=doc_id,
                    kb_id=kid,
                    source="manual",
                )
                db.add(dkb)

        # 更新 documents.knowledge_base_id 為第一個（向下相容）
        if kb_ids and kb_ids[0] in valid_kbs:
            doc.knowledge_base_id = kb_ids[0]

    doc.suggested_kb_id   = None
    doc.suggested_kb_name = None
    await db.commit()
    return {"doc_id": doc_id, "action": action}



@router.patch("/{doc_id}/tags/suggestions")
async def confirm_tag_suggestions(
    doc_id: str,
    body: TagSuggestionAction,
    current_user: CurrentUser = None,
    db: AsyncSession = Depends(get_db),
):
    """
    確認或拒絕 LLM 建議的標籤。
    body: {"action": "confirm", "tag_ids": ["id1", "id2"]}
      → 將選取的 tag_ids 寫入 document_tags（source='llm_suggest'，confidence 從 suggested_tags 取）
      → 清空 suggested_tags
    body: {"action": "reject"}
      → 只清空 suggested_tags
    """
    from models import DocumentTag, Tag

    result = await db.execute(select(Document).where(Document.id == doc_id))
    doc = result.scalar_one_or_none()
    if doc is None:
        raise HTTPException(status_code=404, detail="文件不存在")

    action = body.action
    if action not in ("confirm", "reject"):
        raise HTTPException(status_code=400, detail="action 必須為 confirm 或 reject")

    if action == "confirm" and body.tag_ids:
        # 從 suggested_tags 建立 confidence map
        suggested = doc.suggested_tags or []
        conf_map = {s["tag_id"]: s.get("confidence", 1.0) for s in suggested if s.get("tag_id")}

        # 驗證 tag_ids 都存在
        tags_result = await db.execute(select(Tag).where(Tag.id.in_(body.tag_ids)))
        valid_tags = {t.id for t in tags_result.scalars().all()}

        for tid in body.tag_ids:
            if tid not in valid_tags:
                continue
            # idempotent：ON CONFLICT 用 merge pattern
            existing = await db.execute(
                select(DocumentTag).where(
                    DocumentTag.doc_id == doc_id,
                    DocumentTag.tag_id == tid,
                )
            )
            if existing.scalar_one_or_none() is None:
                dt = DocumentTag(
                    doc_id=doc_id,
                    tag_id=tid,
                    source="llm_suggest",
                    confidence=conf_map.get(tid, 1.0),
                    created_by=current_user.id if current_user else None,
                )
                db.add(dt)

    doc.suggested_tags = []
    await db.commit()
    return {"doc_id": doc_id, "action": action}


@router.patch("/{doc_id}/fields")
async def update_custom_fields(
    doc_id: str,
    fields: dict,
    current_user: CurrentUser = None,
    db: AsyncSession = Depends(get_db),
):
    """更新 custom_fields — 不觸發 re-embedding（只更新 Qdrant payload）"""
    result = await db.execute(select(Document).where(Document.id == doc_id))
    doc = result.scalar_one_or_none()
    if doc is None:
        raise HTTPException(status_code=404, detail="文件不存在")

    updated = {**(doc.custom_fields or {}), **fields}
    doc.custom_fields = updated

    # 同步更新 Qdrant payload（不重新 embed）
    from database import get_qdrant_client
    from sqlalchemy import text

    qdrant = get_qdrant_client()
    chunk_result = await db.execute(
        text("SELECT vector_id FROM chunks WHERE doc_id = :doc_id AND vector_id IS NOT NULL"),
        {"doc_id": doc_id},
    )
    vector_ids = [row[0] for row in chunk_result.fetchall()]
    if vector_ids:
        from config import settings as _cfg
        await qdrant.set_payload(
            collection_name=_cfg.QDRANT_COLLECTION,
            payload={"custom_fields": updated},
            points=vector_ids,
        )

    await db.commit()
    return {"doc_id": doc_id, "custom_fields": updated}


@router.get("/{doc_id}/chunks")
async def get_document_chunks(
    doc_id: str,
    limit: int = Query(default=50, le=200),
    offset: int = Query(default=0, ge=0),
    db: AsyncSession = Depends(get_db),
):
    """取得文件的 chunks 列表（含文字內容）"""
    from models import Chunk

    doc_result = await db.execute(select(Document).where(Document.id == doc_id))
    doc = doc_result.scalar_one_or_none()
    if doc is None:
        raise HTTPException(status_code=404, detail="文件不存在")

    q = (
        select(Chunk)
        .where(Chunk.doc_id == doc_id)
        .order_by(Chunk.chunk_index)
        .limit(limit)
        .offset(offset)
    )
    chunk_result = await db.execute(q)
    chunks = chunk_result.scalars().all()

    return {
        "doc_id": doc_id,
        "title": doc.title,
        "total": doc.chunk_count,
        "chunks": [
            {
                "id": c.id,
                "index": c.chunk_index,
                "content": c.content,
                "page_number": c.page_number,
            }
            for c in chunks
        ],
    }


@router.delete("/{doc_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_document(
    doc_id: str,
    current_user: CurrentUser = None,
    db: AsyncSession = Depends(get_db),
):
    """軟刪除文件（移入垃圾桶，不實際清除向量 / 原檔）"""
    from datetime import datetime, timezone
    result = await db.execute(select(Document).where(Document.id == doc_id))
    doc = result.scalar_one_or_none()
    if doc is None:
        raise HTTPException(status_code=404, detail="文件不存在")

    doc.deleted_at = datetime.now(timezone.utc)
    await db.commit()


@router.post("/{doc_id}/restore", status_code=status.HTTP_200_OK)
async def restore_document(
    doc_id: str,
    current_user: CurrentUser = None,
    db: AsyncSession = Depends(get_db),
):
    """從垃圾桶還原文件"""
    result = await db.execute(select(Document).where(Document.id == doc_id))
    doc = result.scalar_one_or_none()
    if doc is None:
        raise HTTPException(status_code=404, detail="文件不存在")
    if doc.deleted_at is None:
        raise HTTPException(status_code=400, detail="文件不在垃圾桶中")

    doc.deleted_at = None
    await db.commit()
    return {"doc_id": doc_id, "status": "restored"}


@router.delete("/{doc_id}/permanent", status_code=status.HTTP_204_NO_CONTENT)
async def permanent_delete_document(
    doc_id: str,
    current_user: CurrentUser = None,
    db: AsyncSession = Depends(get_db),
):
    """永久刪除文件（含 Qdrant 向量、MinIO 原檔、Neo4j 關係）— Saga 保護"""
    result = await db.execute(select(Document).where(Document.id == doc_id))
    doc = result.scalar_one_or_none()
    if doc is None:
        raise HTTPException(status_code=404, detail="文件不存在")

    from services.saga import saga_transaction
    from database import get_qdrant_client, get_neo4j_driver
    from services.storage import delete_file as minio_delete
    from sqlalchemy import text

    chunk_result = await db.execute(
        text("SELECT vector_id FROM chunks WHERE doc_id = :doc_id AND vector_id IS NOT NULL"),
        {"doc_id": doc_id},
    )
    vector_ids = [row[0] for row in chunk_result.fetchall()]

    with saga_transaction("delete_document", doc_id) as saga:
        if vector_ids:
            from config import settings as _cfg
            qdrant = get_qdrant_client()
            await qdrant.delete(collection_name=_cfg.QDRANT_COLLECTION, points_selector=vector_ids)
        saga.record_step("qdrant")

        driver = get_neo4j_driver()
        async with driver.session(database="neo4j") as neo_session:
            await neo_session.run(
                """
                MATCH (d:Document {id: $doc_id})-[r:MENTIONS]->(e:Entity)
                DELETE r
                WITH e
                WHERE NOT (e)<-[:MENTIONS]-() AND NOT (e)-[:INSTANCE_OF]->()
                DETACH DELETE e
                """,
                doc_id=doc_id,
            )
            await neo_session.run(
                "MATCH (d:Document {id: $doc_id}) DETACH DELETE d", doc_id=doc_id
            )
        saga.record_step("neo4j")

        # 清除外鍵依賴：ontology_review_queue
        await db.execute(
            text("DELETE FROM ontology_review_queue WHERE source_doc_id = :doc_id"),
            {"doc_id": doc_id},
        )

        await db.delete(doc)
        await db.commit()
        saga.record_step("postgres")

        if doc.file_path:
            minio_delete(doc.file_path)
        saga.record_step("minio")


# ── 移動文件至知識庫 ──────────────────────────────────────────

class MoveDocIn(BaseModel):
    knowledge_base_id: str | None  # None = 移出知識庫


# ── 編輯文件中繼資料（標題／描述／註記／圖示）──────────────────

class DocMetaIn(BaseModel):
    title: str | None = None
    description: str | None = None
    notes: str | None = None
    icon: str | None = None


@router.patch("/{doc_id}/meta")
async def update_document_meta(
    doc_id: str,
    body: DocMetaIn,
    current_user: CurrentUser = None,
    db: AsyncSession = Depends(get_db),
):
    """更新文件標題、描述、註記、圖示（不重新分析）"""
    result = await db.execute(select(Document).where(Document.id == doc_id))
    doc = result.scalar_one_or_none()
    if doc is None:
        raise HTTPException(status_code=404, detail="文件不存在")

    if body.title is not None:
        doc.title = body.title.strip() or doc.title

    # 描述、註記、圖示存入 custom_fields
    cf = dict(doc.custom_fields or {})
    if body.description is not None:
        cf["description"] = body.description
    if body.notes is not None:
        cf["notes"] = body.notes
    if body.icon is not None:
        cf["icon"] = body.icon
    doc.custom_fields = cf

    await db.commit()
    return {
        "doc_id": doc_id,
        "title": doc.title,
        "custom_fields": doc.custom_fields,
    }


# ── 重新觸發 AI 分析 ──────────────────────────────────────────

@router.post("/{doc_id}/reanalyze", status_code=status.HTTP_202_ACCEPTED)
async def reanalyze_document(
    doc_id: str,
    current_user: CurrentUser = None,
    db: AsyncSession = Depends(get_db),
):
    """重置文件狀態並重新觸發 Celery 攝取任務"""
    result = await db.execute(select(Document).where(Document.id == doc_id))
    doc = result.scalar_one_or_none()
    if doc is None:
        raise HTTPException(status_code=404, detail="文件不存在")
    if not doc.file_path:
        raise HTTPException(status_code=400, detail="此文件無原始檔案，無法重新分析")

    doc.status = "pending"
    doc.error_message = None
    doc.chunk_count = 0
    await db.commit()

    ingest_document.delay(doc_id)
    logger.info("Reanalyze triggered: %s", doc_id)
    return {"doc_id": doc_id, "status": "pending"}


# ── 下載原始檔案 ──────────────────────────────────────────────

@router.get("/{doc_id}/download")
async def download_document(
    doc_id: str,
    current_user: CurrentUser = None,
    db: AsyncSession = Depends(get_db),
):
    """從 MinIO 取得原始檔案內容（前端用於 SheetJS 渲染）"""
    from fastapi.responses import Response
    from services.storage import download_file as minio_download

    result = await db.execute(select(Document).where(Document.id == doc_id))
    doc = result.scalar_one_or_none()
    if doc is None:
        raise HTTPException(status_code=404, detail="文件不存在")
    if not doc.file_path:
        raise HTTPException(status_code=400, detail="此文件無原始檔案")

    try:
        data = minio_download(doc.file_path)
    except Exception as e:
        logger.error("MinIO download failed %s: %s", doc.file_path, e)
        raise HTTPException(status_code=500, detail="檔案下載失敗")

    content_type_map = {
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "pdf": "application/pdf",
        "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "txt": "text/plain",
        "md": "text/markdown",
        "html": "text/html",
        "csv": "text/csv",
    }
    ct = content_type_map.get(doc.file_type or "", "application/octet-stream")
    return Response(
        content=data,
        media_type=ct,
        headers={"Content-Disposition": f'inline; filename="{doc.title}"'},
    )


@router.patch("/{doc_id}/kb")
async def move_document_to_kb(
    doc_id: str,
    body: MoveDocIn,
    current_user: CurrentUser = None,
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(Document).where(Document.id == doc_id))
    doc = result.scalar_one_or_none()
    if doc is None:
        raise HTTPException(status_code=404, detail="文件不存在")

    if body.knowledge_base_id:
        from models import KnowledgeBase
        kb_result = await db.execute(
            select(KnowledgeBase).where(KnowledgeBase.id == body.knowledge_base_id)
        )
        if kb_result.scalar_one_or_none() is None:
            raise HTTPException(status_code=404, detail="知識庫不存在")

    # 更新舊 FK
    doc.knowledge_base_id = body.knowledge_base_id

    # 同步 document_knowledge_bases 多對多表
    from models import DocumentKnowledgeBase
    from sqlalchemy import delete as sa_delete
    await db.execute(sa_delete(DocumentKnowledgeBase).where(DocumentKnowledgeBase.doc_id == doc_id))
    if body.knowledge_base_id:
        db.add(DocumentKnowledgeBase(doc_id=doc_id, kb_id=body.knowledge_base_id, source="manual"))

    await db.commit()
    return {"doc_id": doc_id, "knowledge_base_id": body.knowledge_base_id}


# ── AI 語意搜尋 ──────────────────────────────────────────────

class DocSearchIn(BaseModel):
    query: str
    kb_id: str | None = None
    top_k: int = 10


class DocSearchResult(BaseModel):
    doc_id: str
    title: str
    file_type: str | None
    status: str
    score: float
    snippet: str
    knowledge_base_id: str | None = None
    knowledge_base_name: str | None = None
    created_at: str


@router.post("/search", response_model=list[DocSearchResult])
async def search_documents_ai(
    body: DocSearchIn,
    current_user: CurrentUser = None,
    db: AsyncSession = Depends(get_db),
):
    """AI 語意搜尋：嵌入查詢 → Qdrant → 聚合返回最相關文件"""
    import httpx
    from database import get_qdrant_client
    from config import settings

    if not body.query.strip():
        return []

    # 1. 嵌入查詢
    try:
        async with httpx.AsyncClient(timeout=60) as client:
            resp = await client.post(
                f"{settings.OLLAMA_BASE_URL}/api/embed",
                json={"model": settings.OLLAMA_EMBED_MODEL, "input": [body.query]},
            )
            resp.raise_for_status()
            query_vec = resp.json().get("embeddings", [[]])[0]
    except Exception as e:
        logger.error("embedding failed: %s", e)
        raise HTTPException(status_code=503, detail="嵌入服務暫時無法使用")

    # 2. 如果指定 KB，先取該 KB 的 doc_id 列表
    search_filter = None
    if body.kb_id:
        kb_docs_result = await db.execute(
            select(Document.id).where(Document.knowledge_base_id == body.kb_id)
        )
        kb_doc_ids = [r[0] for r in kb_docs_result.all()]
        if not kb_doc_ids:
            return []
        from qdrant_client.models import Filter, FieldCondition, MatchAny
        search_filter = Filter(
            must=[FieldCondition(key="doc_id", match=MatchAny(any=kb_doc_ids))]
        )

    # 3. Qdrant 向量搜尋
    qdrant = get_qdrant_client()
    hits = await qdrant.search(
        collection_name=settings.QDRANT_COLLECTION,
        query_vector=query_vec,
        limit=body.top_k * 5,
        query_filter=search_filter,
        with_payload=True,
    )

    # 4. 聚合：每個 doc_id 保留最高分
    seen: dict[str, dict] = {}
    for hit in hits:
        doc_id = hit.payload.get("doc_id")
        if doc_id and (doc_id not in seen or hit.score > seen[doc_id]["score"]):
            seen[doc_id] = {
                "score": hit.score,
                "snippet": (hit.payload.get("content") or "")[:200],
            }

    if not seen:
        return []

    # 5. 取 Document 資訊
    docs_result = await db.execute(
        select(Document).where(Document.id.in_(list(seen.keys())))
    )
    docs_map = {d.id: d for d in docs_result.scalars().all()}

    # load KB names
    from models import KnowledgeBase
    kb_ids = {d.knowledge_base_id for d in docs_map.values() if d.knowledge_base_id}
    kb_name_map: dict[str, str] = {}
    if kb_ids:
        kb_result = await db.execute(select(KnowledgeBase).where(KnowledgeBase.id.in_(kb_ids)))
        kb_name_map = {kb.id: kb.name for kb in kb_result.scalars().all()}

    # 6. 排序並返回
    results = []
    for doc_id, info in sorted(seen.items(), key=lambda x: -x[1]["score"]):
        if doc_id not in docs_map:
            continue
        d = docs_map[doc_id]
        results.append(DocSearchResult(
            doc_id=doc_id,
            title=d.title,
            file_type=d.file_type,
            status=d.status,
            score=round(info["score"], 4),
            snippet=info["snippet"],
            knowledge_base_id=d.knowledge_base_id,
            knowledge_base_name=kb_name_map.get(d.knowledge_base_id) if d.knowledge_base_id else None,
            created_at=d.created_at.isoformat(),
        ))

    return results[: body.top_k]

